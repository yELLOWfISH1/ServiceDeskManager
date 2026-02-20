"""
Active Directory Tab - PC Management, Movement, and User Search
"""
import configparser
import csv
from datetime import datetime
from pathlib import Path
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, 
    QComboBox, QMessageBox, QTextEdit, QFileDialog, QFrame
)
from PySide6.QtCore import Qt
from openpyxl import load_workbook
from .ldap_manager import LDAPManager


class ADTab(QWidget):
    """Active Directory operations tab"""
    
    def __init__(self, settings_tab=None):
        super().__init__()
        self.settings_tab = settings_tab
        self.ldap_config = self.load_ad_config()
        self.ou_list = self.parse_ou_list()
        self.ldap_manager = None
        self.setup_ui()
    
    def load_ad_config(self):
        """Load AD configuration from config.ini"""
        config = configparser.ConfigParser()
        config_path = Path(__file__).parent / "config.ini"
        if config_path.exists():
            config.read(config_path)
            return dict(config.items('AD')) if 'AD' in config else {}
        return {}
    
    def parse_ou_list(self):
        """Parse OUs from config - returns dict of FriendlyName -> OU_DN"""
        ou_string = self.ldap_config.get('organizational_units', '')
        ou_dict = {}
        
        for item in ou_string.split('|'):
            item = item.strip()
            if ':' in item:
                friendly_name, ou_dn = item.split(':', 1)
                ou_dict[friendly_name.strip()] = ou_dn.strip()
        
        return ou_dict

    def _get_ad_log_file_path(self):
        """Resolve AD audit log file path from config (default: logs/ad_audit.csv)."""
        log_folder_setting = self.ldap_config.get('log_folder', 'logs').strip() or 'logs'
        folder_path = Path(log_folder_setting)
        if not folder_path.is_absolute():
            folder_path = Path(__file__).parent.parent / folder_path
        folder_path.mkdir(parents=True, exist_ok=True)
        return folder_path / "ad_audit.csv"

    def _write_ad_audit_log(self, action, hostname, status, message, target_ou=''):
        """Write AD action audit entry (delete/move/description only)."""
        try:
            username, _ = self.get_credentials()
            actor = username if username else "unknown"
            log_file = self._get_ad_log_file_path()
            file_exists = log_file.exists()

            with open(log_file, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow([
                        "timestamp",
                        "actor",
                        "action",
                        "hostname",
                        "target_ou",
                        "status",
                        "message",
                    ])
                writer.writerow([
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    actor,
                    action,
                    hostname,
                    target_ou,
                    status,
                    message,
                ])
        except Exception:
            pass

    def setup_ui(self):
        """Create the AD tab UI"""
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)
        
        # ===== DELETE PC SECTION =====
        delete_title = QLabel("Delete Computer from AD")
        delete_title.setObjectName("section-title")
        delete_title.setStyleSheet("color: #e74c3c;")
        main_layout.addWidget(delete_title)
        
        delete_container = QFrame()
        delete_layout = QVBoxLayout()
        delete_layout.setContentsMargins(0, 0, 0, 0)
        delete_layout.setSpacing(8)
        
        # Hostname input
        delete_input_layout = QHBoxLayout()
        self.delete_hostname = QLineEdit()
        self.delete_hostname.setPlaceholderText("Enter computer hostname (e.g., COMPUTER-01)")
        delete_input_layout.addWidget(self.delete_hostname)
        
        delete_btn = QPushButton("🗑️ Delete PC")
        delete_btn.setMinimumHeight(32)
        delete_btn.setMaximumWidth(140)
        delete_btn.clicked.connect(self.delete_pc)
        delete_input_layout.addWidget(delete_btn)
        delete_layout.addLayout(delete_input_layout)
        
        # Status message
        self.delete_status = QLabel("")
        self.delete_status.setStyleSheet("font-size: 10px; color: #888888;")
        self.delete_status.setWordWrap(True)
        delete_layout.addWidget(self.delete_status)
        
        delete_container.setLayout(delete_layout)
        main_layout.addWidget(delete_container)
        
        # Separator
        separator1 = QFrame()
        separator1.setFrameShape(QFrame.HLine)
        separator1.setFrameShadow(QFrame.Sunken)
        main_layout.addWidget(separator1)
        
        # ===== MOVE PC SECTION =====
        move_title = QLabel("Move Computer to OU")
        move_title.setObjectName("section-title")
        move_title.setStyleSheet("color: #3498db;")
        main_layout.addWidget(move_title)
        
        move_container = QFrame()
        move_layout = QVBoxLayout()
        move_layout.setContentsMargins(0, 0, 0, 0)
        move_layout.setSpacing(8)
        
        # Single Move
        single_label = QLabel("Single Move")
        single_label.setStyleSheet("color: #b0b0b0; font-size: 11px; font-weight: bold;")
        move_layout.addWidget(single_label)
        
        single_row = QHBoxLayout()
        self.move_hostname = QLineEdit()
        self.move_hostname.setPlaceholderText("Hostname")
        single_row.addWidget(self.move_hostname)
        
        self.ou_dropdown = QComboBox()
        self.ou_dropdown.addItems(self.ou_list.keys() if self.ou_list else ["No OUs configured"])
        single_row.addWidget(self.ou_dropdown)
        
        move_btn = QPushButton("Move")
        move_btn.setMinimumHeight(32)
        move_btn.setMaximumWidth(100)
        move_btn.clicked.connect(self.move_pc)
        single_row.addWidget(move_btn)
        move_layout.addLayout(single_row)
        
        # Bulk Move
        bulk_label = QLabel("Bulk Move from Excel")
        bulk_label.setStyleSheet("color: #b0b0b0; font-size: 11px; font-weight: bold; margin-top: 12px;")
        move_layout.addWidget(bulk_label)
        
        bulk_file_row = QHBoxLayout()
        self.bulk_file_label = QLineEdit()
        self.bulk_file_label.setReadOnly(True)
        self.bulk_file_label.setPlaceholderText("No file selected")
        bulk_file_row.addWidget(self.bulk_file_label)
        
        browse_btn = QPushButton("Browse")
        browse_btn.setMinimumHeight(32)
        browse_btn.setMaximumWidth(100)
        browse_btn.clicked.connect(self.browse_excel)
        bulk_file_row.addWidget(browse_btn)
        move_layout.addLayout(bulk_file_row)
        
        bulk_ou_row = QHBoxLayout()
        ou_label = QLabel("Target OU:")
        ou_label.setStyleSheet("color: #b0b0b0;")
        ou_label.setMaximumWidth(70)
        bulk_ou_row.addWidget(ou_label)
        
        self.bulk_ou_dropdown = QComboBox()
        self.bulk_ou_dropdown.addItems(self.ou_list.keys() if self.ou_list else ["No OUs configured"])
        bulk_ou_row.addWidget(self.bulk_ou_dropdown)
        
        bulk_move_btn = QPushButton("Bulk Move")
        bulk_move_btn.setMinimumHeight(32)
        bulk_move_btn.setMaximumWidth(120)
        bulk_move_btn.clicked.connect(self.bulk_move_pc)
        bulk_ou_row.addWidget(bulk_move_btn)
        bulk_ou_row.addStretch()
        move_layout.addLayout(bulk_ou_row)
        
        # Move status message
        self.move_status = QLabel("")
        self.move_status.setStyleSheet("font-size: 10px; color: #888888;")
        self.move_status.setWordWrap(True)
        move_layout.addWidget(self.move_status)
        
        move_container.setLayout(move_layout)
        main_layout.addWidget(move_container)
        
        # Separator
        separator2 = QFrame()
        separator2.setFrameShape(QFrame.HLine)
        separator2.setFrameShadow(QFrame.Sunken)
        main_layout.addWidget(separator2)

        # ===== EDIT DESCRIPTION SECTION =====
        desc_title = QLabel("Edit Computer Description")
        desc_title.setObjectName("section-title")
        desc_title.setStyleSheet("color: #f39c12;")
        main_layout.addWidget(desc_title)

        desc_container = QFrame()
        desc_layout = QVBoxLayout()
        desc_layout.setContentsMargins(0, 0, 0, 0)
        desc_layout.setSpacing(8)

        desc_info = QLabel("Requires admin credentials from Settings tab")
        desc_info.setStyleSheet("color: #999999; font-size: 10px;")
        desc_layout.addWidget(desc_info)

        desc_row1 = QHBoxLayout()
        self.desc_hostname = QLineEdit()
        self.desc_hostname.setPlaceholderText("Hostname (e.g., COMPUTER-01)")
        desc_row1.addWidget(self.desc_hostname)
        desc_layout.addLayout(desc_row1)

        desc_row2 = QHBoxLayout()
        self.desc_value = QLineEdit()
        self.desc_value.setPlaceholderText("New description (leave blank to clear)")
        desc_row2.addWidget(self.desc_value)

        desc_update_btn = QPushButton("Update Description")
        desc_update_btn.setMinimumHeight(32)
        desc_update_btn.setMaximumWidth(170)
        desc_update_btn.clicked.connect(self.update_pc_description)
        desc_row2.addWidget(desc_update_btn)
        desc_layout.addLayout(desc_row2)

        self.description_status = QLabel("")
        self.description_status.setStyleSheet("font-size: 10px; color: #888888;")
        self.description_status.setWordWrap(True)
        desc_layout.addWidget(self.description_status)

        desc_container.setLayout(desc_layout)
        main_layout.addWidget(desc_container)

        # Separator
        separator3 = QFrame()
        separator3.setFrameShape(QFrame.HLine)
        separator3.setFrameShadow(QFrame.Sunken)
        main_layout.addWidget(separator3)
        
        # ===== USER SEARCH SECTION =====
        user_title = QLabel("Search User Account")
        user_title.setObjectName("section-title")
        user_title.setStyleSheet("color: #2ecc71;")
        main_layout.addWidget(user_title)
        
        user_container = QFrame()
        user_layout = QVBoxLayout()
        user_layout.setContentsMargins(0, 0, 0, 0)
        user_layout.setSpacing(8)
        
        info_label = QLabel("Enter full or partial name/user ID (e.g., Jack) to list all matches")
        info_label.setStyleSheet("color: #999999; font-size: 10px;")
        user_layout.addWidget(info_label)
        
        search_row = QHBoxLayout()
        self.user_search = QLineEdit()
        self.user_search.setPlaceholderText("e.g., John Smith or jsmith")
        search_row.addWidget(self.user_search)
        
        search_btn = QPushButton("Search")
        search_btn.setMinimumHeight(32)
        search_btn.setMaximumWidth(100)
        search_btn.clicked.connect(self.search_user)
        search_row.addWidget(search_btn)
        user_layout.addLayout(search_row)
        
        # Results display
        self.user_result = QTextEdit()
        self.user_result.setReadOnly(True)
        self.user_result.setMinimumHeight(100)
        self.user_result.setStyleSheet("""
            QTextEdit {
                background-color: #0f1419;
                color: #2ecc71;
                border: 1px solid #333333;
                border-radius: 4px;
                padding: 8px;
                font-family: Consolas, monospace;
                font-size: 10px;
            }
        """)
        user_layout.addWidget(self.user_result)
        
        user_container.setLayout(user_layout)
        main_layout.addWidget(user_container)
        
        main_layout.addStretch()
        self.setLayout(main_layout)

    def test_ad_connection(self, show_dialog=True):
        """Test AD connectivity using configured server/base DN and optional credentials."""
        username, password = self.get_credentials()

        manager = LDAPManager(
            self.ldap_config.get('ldap_server', 'localhost'),
            int(self.ldap_config.get('ldap_port', '389')),
            self.ldap_config.get('ldap_base_dn', ''),
            username or '',
            password or ''
        )

        ok = manager.connect()
        manager.disconnect()

        if show_dialog:
            if ok:
                auth_mode = "configured admin credentials" if username and password else "current Windows credentials"
                QMessageBox.information(
                    self,
                    "AD Connection Test",
                    f"✅ AD connection successful using {auth_mode}."
                )
            else:
                QMessageBox.warning(
                    self,
                    "AD Connection Test",
                    "❌ AD connection failed.\n\nCheck Settings credentials and src/config.ini AD values."
                )

        return ok
    
    def get_credentials(self):
        """Get credentials from settings tab"""
        if self.settings_tab:
            return self.settings_tab.get_credentials()
        return None, None
    
    def delete_pc(self):
        """Delete a PC from AD"""
        hostname = self.delete_hostname.text().strip()
        if not hostname:
            self.delete_status.setText("❌ Please enter a hostname")
            self.delete_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("DELETE", "", "FAILED", "Hostname missing")
            return
        
        username, password = self.get_credentials()
        if not username or not password:
            QMessageBox.warning(self, "Credentials Required", 
                "Please set admin credentials in Settings tab first.")
            self._write_ad_audit_log("DELETE", hostname, "FAILED", "Missing admin credentials")
            return
        
        # Initialize LDAP manager
        self.ldap_manager = LDAPManager(
            self.ldap_config.get('ldap_server', 'localhost'),
            int(self.ldap_config.get('ldap_port', '389')),
            self.ldap_config.get('ldap_base_dn', ''),
            username,
            password
        )
        
        # Connect and get PC info
        if not self.ldap_manager.connect():
            self.delete_status.setText("❌ Failed to connect to AD")
            self.delete_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("DELETE", hostname, "FAILED", "Failed to connect to AD")
            return
        
        pc_info = self.ldap_manager.get_computer_info(hostname)
        self.ldap_manager.disconnect()
        
        if not pc_info:
            self.delete_status.setText(f"❌ Computer '{hostname}' not found in AD")
            self.delete_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("DELETE", hostname, "FAILED", "Computer not found in AD")
            return
        
        # Show confirmation with PC details
        description = pc_info.get('description', 'No description')
        ou = pc_info.get('ou', 'Unknown OU')
        
        reply = QMessageBox.warning(self, "Confirm PC Deletion",
            f"⚠️ Delete Computer?\n\n"
            f"Hostname: {hostname}\n"
            f"Description: {description}\n"
            f"Current OU: {ou}\n\n"
            f"Please verify in ServiceNow first!\n\n"
            "Are you sure you want to continue?",
            QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            # Reconnect for deletion
            self.ldap_manager = LDAPManager(
                self.ldap_config.get('ldap_server', 'localhost'),
                int(self.ldap_config.get('ldap_port', '389')),
                self.ldap_config.get('ldap_base_dn', ''),
                username,
                password
            )
            
            if not self.ldap_manager.connect():
                self.delete_status.setText("❌ Failed to connect to AD for deletion")
                self.delete_status.setStyleSheet("font-size: 9px; color: #f44336;")
                self._write_ad_audit_log("DELETE", hostname, "FAILED", "Failed to connect for deletion")
                return
            
            success, message = self.ldap_manager.delete_computer(hostname)
            self.ldap_manager.disconnect()
            
            if success:
                self.delete_status.setText(f"✅ {message}")
                self.delete_status.setStyleSheet("font-size: 9px; color: #4CAF50;")
                self.delete_hostname.clear()
                self._write_ad_audit_log("DELETE", hostname, "SUCCESS", message)
            else:
                self.delete_status.setText(f"❌ {message}")
                self.delete_status.setStyleSheet("font-size: 9px; color: #f44336;")
                self._write_ad_audit_log("DELETE", hostname, "FAILED", message)
    
    def move_pc(self):
        """Move a single PC to an OU"""
        hostname = self.move_hostname.text().strip()
        friendly_name = self.ou_dropdown.currentText()
        
        if not hostname:
            self.move_status.setText("❌ Please enter a hostname")
            self.move_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("MOVE", "", "FAILED", "Hostname missing")
            return
        
        if not self.ou_list or friendly_name not in self.ou_list:
            self.move_status.setText("❌ Invalid OU selected")
            self.move_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("MOVE", hostname, "FAILED", "Invalid OU selected", friendly_name)
            return
        
        target_ou = self.ou_list[friendly_name]
        
        username, password = self.get_credentials()
        if not username or not password:
            QMessageBox.warning(self, "Credentials Required",
                "Please set admin credentials in Settings tab first.")
            self._write_ad_audit_log("MOVE", hostname, "FAILED", "Missing admin credentials", friendly_name)
            return
        
        reply = QMessageBox.information(self, "Confirm Move",
            f"Moving '{hostname}' to:\n\n{friendly_name}\n\nContinue?",
            QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            self.ldap_manager = LDAPManager(
                self.ldap_config.get('ldap_server', 'localhost'),
                int(self.ldap_config.get('ldap_port', '389')),
                self.ldap_config.get('ldap_base_dn', ''),
                username,
                password
            )
            
            if not self.ldap_manager.connect():
                self.move_status.setText("❌ Failed to connect to AD")
                self.move_status.setStyleSheet("font-size: 9px; color: #f44336;")
                self._write_ad_audit_log("MOVE", hostname, "FAILED", "Failed to connect to AD", friendly_name)
                return
            
            success, message = self.ldap_manager.move_computer(hostname, target_ou)
            self.ldap_manager.disconnect()
            
            if success:
                self.move_status.setText(f"✅ {message}")
                self.move_status.setStyleSheet("font-size: 9px; color: #4CAF50;")
                self.move_hostname.clear()
                self._write_ad_audit_log("MOVE", hostname, "SUCCESS", message, friendly_name)
            else:
                self.move_status.setText(f"❌ {message}")
                self.move_status.setStyleSheet("font-size: 9px; color: #f44336;")
                self._write_ad_audit_log("MOVE", hostname, "FAILED", message, friendly_name)
    
    def browse_excel(self):
        """Browse and select Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", 
            "Excel Files (*.xlsx *.xls);;All Files (*)")
        if file_path:
            self.bulk_file_label.setText(file_path)
    
    def bulk_move_pc(self):
        """Bulk move PCs from Excel file"""
        file_path = self.bulk_file_label.text()
        friendly_name = self.bulk_ou_dropdown.currentText()
        
        if not file_path:
            self.move_status.setText("❌ Please select an Excel file")
            self.move_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("BULK_MOVE", "", "FAILED", "No Excel file selected")
            return
        
        if not self.ou_list or friendly_name not in self.ou_list:
            self.move_status.setText("❌ Invalid OU selected")
            self.move_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("BULK_MOVE", "", "FAILED", "Invalid OU selected", friendly_name)
            return
        
        target_ou = self.ou_list[friendly_name]
        
        username, password = self.get_credentials()
        if not username or not password:
            QMessageBox.warning(self, "Credentials Required",
                "Please set admin credentials in Settings tab first.")
            self._write_ad_audit_log("BULK_MOVE", "", "FAILED", "Missing admin credentials", friendly_name)
            return
        
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            hostnames = [str(row[0].value).strip() for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1) if row[0].value]
            
            if not hostnames:
                self.move_status.setText("❌ No hostnames found in Excel file")
                self.move_status.setStyleSheet("font-size: 9px; color: #f44336;")
                self._write_ad_audit_log("BULK_MOVE", "", "FAILED", "No hostnames found in Excel", friendly_name)
                return
            
            hostname_list = ", ".join(hostnames[:10]) + ("..." if len(hostnames) > 10 else "")
            reply = QMessageBox.information(self, "Confirm Bulk Move",
                f"Moving {len(hostnames)} computer(s):\n{hostname_list}\n\nTo OU:\n{friendly_name}\n\nContinue?",
                QMessageBox.Yes | QMessageBox.No)
            
            if reply == QMessageBox.Yes:
                self.ldap_manager = LDAPManager(
                    self.ldap_config.get('ldap_server', 'localhost'),
                    int(self.ldap_config.get('ldap_port', '389')),
                    self.ldap_config.get('ldap_base_dn', ''),
                    username,
                    password
                )
                
                if not self.ldap_manager.connect():
                    self.move_status.setText("❌ Failed to connect to AD")
                    self.move_status.setStyleSheet("font-size: 9px; color: #f44336;")
                    self._write_ad_audit_log("BULK_MOVE", "", "FAILED", "Failed to connect to AD", friendly_name)
                    return
                
                # Move each computer
                success_count = 0
                for hostname in hostnames:
                    success, msg = self.ldap_manager.move_computer(hostname, target_ou)
                    if success:
                        success_count += 1
                        self._write_ad_audit_log("MOVE", hostname, "SUCCESS", msg, friendly_name)
                    else:
                        self._write_ad_audit_log("MOVE", hostname, "FAILED", msg, friendly_name)
                
                self.ldap_manager.disconnect()
                
                self.move_status.setText(f"✅ Bulk move complete: {success_count}/{len(hostnames)} computers moved to {friendly_name}")
                self.move_status.setStyleSheet("font-size: 9px; color: #4CAF50;")
                self.bulk_file_label.clear()
                self._write_ad_audit_log(
                    "BULK_MOVE",
                    "",
                    "SUCCESS" if success_count == len(hostnames) else "PARTIAL",
                    f"Bulk move complete: {success_count}/{len(hostnames)}",
                    friendly_name
                )
        
        except Exception as e:
            self.move_status.setText(f"❌ Error reading file: {str(e)}")
            self.move_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("BULK_MOVE", "", "FAILED", f"Error reading file: {str(e)}", friendly_name)

    def update_pc_description(self):
        """Update a computer object's description in AD (admin required)."""
        hostname = self.desc_hostname.text().strip()
        description = self.desc_value.text().strip()

        if not hostname:
            self.description_status.setText("❌ Please enter a hostname")
            self.description_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("UPDATE_DESCRIPTION", "", "FAILED", "Hostname missing")
            return

        username, password = self.get_credentials()
        if not username or not password:
            QMessageBox.warning(self, "Credentials Required",
                "Please set admin credentials in Settings tab first.")
            self._write_ad_audit_log("UPDATE_DESCRIPTION", hostname, "FAILED", "Missing admin credentials")
            return

        self.ldap_manager = LDAPManager(
            self.ldap_config.get('ldap_server', 'localhost'),
            int(self.ldap_config.get('ldap_port', '389')),
            self.ldap_config.get('ldap_base_dn', ''),
            username,
            password
        )

        if not self.ldap_manager.connect():
            self.description_status.setText("❌ Failed to connect to AD")
            self.description_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("UPDATE_DESCRIPTION", hostname, "FAILED", "Failed to connect to AD")
            return

        success, message = self.ldap_manager.update_computer_description(hostname, description)
        self.ldap_manager.disconnect()

        if success:
            self.description_status.setText(f"✅ {message}")
            self.description_status.setStyleSheet("font-size: 9px; color: #4CAF50;")
            self.desc_hostname.clear()
            self.desc_value.clear()
            self._write_ad_audit_log("UPDATE_DESCRIPTION", hostname, "SUCCESS", message)
        else:
            self.description_status.setText(f"❌ {message}")
            self.description_status.setStyleSheet("font-size: 9px; color: #f44336;")
            self._write_ad_audit_log("UPDATE_DESCRIPTION", hostname, "FAILED", message)
    
    def search_user(self):
        """Search for user's OU"""
        search_term = self.user_search.text().strip()
        
        if not search_term:
            self.user_result.setText("❌ Please enter a name or user ID")
            return
        
        # User search doesn't require credentials per user request
        # Initialize LDAP manager with anonymous bind (empty credentials)
        self.ldap_manager = LDAPManager(
            self.ldap_config.get('ldap_server', 'localhost'),
            int(self.ldap_config.get('ldap_port', '389')),
            self.ldap_config.get('ldap_base_dn', ''),
            '',  # Empty username for anonymous bind
            ''   # Empty password for anonymous bind
        )
        
        try:
            if not self.ldap_manager.connect():
                self.user_result.setText("❌ Failed to connect to AD server")
                return

            users = self.ldap_manager.search_users(search_term)
            self.ldap_manager.disconnect()

            if not users:
                self.user_result.setText(f"❌ No users found for '{search_term}'\n\nTip: Try partial search like first name, last name, or userid fragment")
                return

            lines = [f"Found {len(users)} user(s):", ""]
            for i, user_info in enumerate(users, start=1):
                lines.append(f"[{i}] Display Name: {user_info['displayName']}")
                lines.append(f"    User ID: {user_info['sAMAccountName']}")
                lines.append(f"    OU: {user_info['ou_windows']}")
                lines.append("")

            lines.append("[Easy Copy: Highlight and Ctrl+C above]")
            self.user_result.setText("\n".join(lines))
            self.user_search.clear()

        except Exception as e:
            self.user_result.setText(f"❌ Error searching user: {str(e)}")
